import os
import openpyxl
import csv
protein_data = []
row_counter = 1
def prompt():
    print('Select an option from the following list by typing in the number:')
    print('1. Organize Sadygov Rate Constants only.')
    print('2. Calculate rate constants using Prism.')
    print('3. To run both.')
    print('4. To calculate FSR from two time points.')
    selection = input('Enter your choice: ').strip()
    while check_selection(selection) == False:
        print('\n')
        print('This is an invalid input.')
        print('Select an option from the following list by typing in the number:')
        print('1. Organize Sadygov Rate Constants only.')
        print('2. Calculate rate constants using Prism.')
        print('3. To run both.')
        print('4. To calculate FSR from two time points.')
        selection = input('Enter your choice: ').strip()
    return selection

def check_selection(selection):
    if selection == '1' or selection == '2' or selection == '3' or selection =='4':
        return True
    else:
        return False
def prompt_sadygov():
    print('\n')
    print('Please make sure the folder path that contains your data does not contain any spaces. I.e C:\Kasumov lab\Andrew\Data_Folder will not work because there is a space in "Kasumov lab".')
    print('Please enter the path of the folder that contains your files. I.e C:\Kasumov_lab\Andrew\Data_Folder')
    folder_name = input('Folder Path: ').strip()
    folder_name = check_directory(folder_name)
    
    return folder_name

def check_directory_sadygov(folder_name):
    while os.path.isdir(folder_name) == False:
        print('\n')
        print('This directory either does not exist or is mispelled.')
        print('Please enter the path of the folder that contains your files. I.e C:\Kasumov_lab\Andrew\Data_Folder')
        folder_name = input('Folder Path: ').strip()
    return folder_name
def read_file_sadygov(folder_name):
    data_list = [['Protein','Rate Constant']]
    for file_name in os.listdir(folder_name):
        file_name_split = file_name.split('.')
        if len(file_name_split) < 3:
            pass
        elif file_name_split[1] == 'RateConst':
            file_path = os.path.sep.join([folder_name, file_name])
            with open(file_path) as file:
                for line in file:
                    line = line.split(',')
                    if line[0].strip() == 'MeanRateConst/CorrCutOf':
                        split = file_name_split[0].split('_')
                        data_list.append([split[0],line[1]])
                    else:
                        pass
                    
        else:
            pass
    return data_list
            
def write_file_sadygov(data_list, folder_name):
    import csv
    file_path_all = os.path.sep.join([folder_name, 'All_Protein_RateConstant.csv'])
    with open(file_path_all, 'w', newline = '') as final_file:
        a = csv.writer(final_file, delimiter = ',')
        a.writerows(data_list)
def prompt_tl():
    print('\n')
    print('Please make sure the folder path that contains your data does not contain any spaces. I.e C:\Kasumov lab\Andrew\Data_Folder will not work because there is a space in "Kasumov lab".')
    print('Please enter the path of the folder that contains your files. I.e C:\Kasumov_lab\Andrew\Data_Folder')
    folder_name = input('Folder Path: ').strip()
    folder_name = check_directory(folder_name)
    

    print('\n')
    print('Please select which isotopomer you would like to use to calculate the rate constant.')
    print('1. I1')
    print('2. I2')
    print('3. I3')
    print('4. I4')
    print('5. I5')
    prism_input = input('Enter your choice: ').strip()
    while check_prism_input(prism_input) == False:
        print('This is an invalid input, please try again. \n')
        print('\n')
        print('Please select which isotopomer you would like to use to calculate the rate constant.')
        print('1. I1')
        print('2. I2')
        print('3. I3')
        print('4. I4')
        print('5. I5')
        prism_input = input('Enter your choice: ').strip()
        
    print('\n')        
    time_points = input('Please enter the timepoints separated by commas without any letters, I.e "0,0,8,8,24,....": ').strip()
    while check_time_points(time_points,folder_name) == False:
        print('\n')
        print('The time points you have entered do not match the time points listed in the folder.')
        print('Please try again.')
        time_points = input('Please enter the timepoints separated by commas without any letters, I.e "0,0,8,8,24,....": ').strip()
        

    print('\n')
    unique = input('Please enter "yes" to use only unique peptides or "no" to use all peptides: ').strip().lower()
    while check_input(unique) == False:
        print('\n')
        print('This is an invalid input.')
        unique = input('Please enter "yes" to use only unique peptides or "no" to use all peptides: ').strip().lower()
    return folder_name, time_points, unique, prism_input

def check_prism_input(prism_input):
    if prism_input == '1' or prism_input == '2' or prism_input == '3' or prism_input == '4' or prism_input == '5':
        return True
    else:
        return False
def check_time_points(time_points,folder_name):
    counter = 0
    for file in os.listdir(folder_name):
        file = file.split('.')
        try:
            if file[1] == 'Proteins':
                counter +=1
            else:
                pass
        except IndexError:
            pass
    if len(time_points.split(',')) == counter:
        return True
    else:
        return False
def check_directory(folder_name):
    while os.path.isdir(folder_name) == False or ' ' in folder_name:
        print('\n')
        print('This directory either does not exist, is mispelled, or contains a space.')
        print('Please enter the path of the folder that contains your files. I.e C:\Kasumov_lab\Andrew\Data_Folder')
        folder_name = input('Folder Path: ').strip()
    return folder_name
def check_input(unique):
    if unique == 'no' or unique == 'yes':
        return True
    else:
        return False
        
def read_data_A4(file, file_name, file_path, folder_name, time_points, prism_input):
    time_point_list = ['']
    peptides_list = ['']
    peptides_dict = {} #Need to add peptides to this
    rowq_counter = 1 #Counter for which row I am on
    with open(file_path) as file:
        for row in file:
            row = row.split(',')
            if rowq_counter == 1 or rowq_counter == 2:
                pass
            elif rowq_counter == 3:#Making list of Timepoints
                (peptides_dict, time_point_list) = remove_blanks_from_list(row, time_point_list, peptides_dict)
            elif rowq_counter ==4:
                data_locations = []
                for position,value in enumerate(row):
                    if value.strip() == 'I0' or value.strip() == 'I1' or value.strip() == 'I2' or value.strip() == 'I3' or value.strip() == 'I4'or value.strip() == 'I5':
                        data_locations.append(position)
                    else:
                        pass
            else:
                row[0] = row[0] + str(rowq_counter)
                peptides_list.append(row[0])
                peptides_dict[row[0]] = []
                data_list = []
                counter_random = 1
                for i in data_locations:
                    if test_for_number(row,i)==True:
                        data_list.append(row[i])
                    else:
                        data_list.append(0)
                    if counter_random == 6:
                        counter_random = 0
                        (I0,I1,I2,I3,I4,I5) = time_names_fxn(data_list)
                        peptides_dict = test_for_zeros(I0,I1,I2,I3,I4,I5,peptides_dict,row)
                        data_list = []
                    else:
                        pass
                    counter_random +=1
            rowq_counter += 1
    if len(peptides_list) == 2:
        pass
    else:
        write_data_main(file_name, peptides_list, time_point_list, peptides_dict, folder_name, time_points, prism_input)

def read_data_A2(file, file_name, file_path, folder_name, time_points, prism_input):
    time_point_list = ['']
    peptides_list = ['']
    peptides_dict = {} #Need to add peptides to this
    rowq_counter = 1 #Counter for which row I am on
    with open(file_path) as file:
        for row in file:
            row = row.split(',')
            if rowq_counter == 1:#Making list of Timepoints
                (peptides_dict, time_point_list) = remove_blanks_from_list(row, time_point_list, peptides_dict)
            elif rowq_counter == 2:
                data_locations = []
                for position,value in enumerate(row):
                    if value.strip() == 'I0' or value.strip() == 'I1' or value.strip() == 'I2' or value.strip() == 'I3' or value.strip() == 'I4'or value.strip() == 'I5':
                        data_locations.append(position)
                    else:
                        pass
            else:
                row[0] = row[0] + str(rowq_counter)
                peptides_list.append(row[0])
                peptides_dict[row[0]] = []
                data_list = []
                counter_random = 1
                for i in data_locations:
                    if test_for_number(row,i)==True:
                        data_list.append(row[i])
                    else:
                        data_list.append(0)
                    if counter_random == 6:
                        counter_random = 0
                        (I0,I1,I2,I3,I4,I5) = time_names_fxn(data_list)
                        peptides_dict = test_for_zeros(I0,I1,I2,I3,I4,I5,peptides_dict,row)
                        data_list = []
                    else:
                        pass
                    counter_random +=1
            rowq_counter += 1
    if len(peptides_list) == 2:
        pass
    else:
        write_data_main(file_name, peptides_list, time_point_list, peptides_dict, folder_name, time_points, prism_input)

def read_data_A4_unique(file, file_name, file_path, folder_name, time_points, prism_input):
    time_point_list = ['']
    peptides_list = ['']
    peptides_dict = {} #Need to add peptides to this
    rowq_counter = 1 #Counter for which row I am on
    with open(file_path) as file:
        for row in file:
            row = row.split(',')
            if rowq_counter == 1 or rowq_counter == 2:
                pass
            elif rowq_counter == 3:#Making list of Timepoints
                (peptides_dict, time_point_list) = remove_blanks_from_list(row, time_point_list, peptides_dict)
            elif rowq_counter ==4:
                data_locations = []
                for position,value in enumerate(row):
                    if value.strip() == 'I0' or value.strip() == 'I1' or value.strip() == 'I2' or value.strip() == 'I3' or value.strip() == 'I4'or value.strip() == 'I5':
                        data_locations.append(position)
                    else:
                        pass
            else:
                row[0] = row[0] + str(rowq_counter)
                peptides_list.append(row[0])
                peptides_dict[row[0]] = []
                data_list = []
                counter_random = 1
                for i in data_locations:
                    if test_for_number(row,i)==True:
                        data_list.append(row[i])
                    else:
                        data_list.append(0)
                    if counter_random == 6:
                        counter_random = 0
                        (I0,I1,I2,I3,I4,I5) = time_names_fxn(data_list)
                        peptides_dict = test_for_zeros(I0,I1,I2,I3,I4,I5,peptides_dict,row)
                        data_list = []
                    else:
                        pass
                    counter_random +=1
            rowq_counter += 1
    if len(peptides_list) == 2:
        pass
    else:
        write_data_main(file_name, peptides_list, time_point_list, peptides_dict, folder_name, time_points, prism_input)

def read_data_A2_unique(file, file_name, file_path, folder_name, time_points, prism_input):
    time_point_list = ['']
    peptides_list = ['']
    peptides_dict = {} #Need to add peptides to this
    rowq_counter = 1 #Counter for which row I am on
    with open(file_path) as file:
        for row in file:
            row = row.split(',')
            if rowq_counter == 1:#Making list of Timepoints
                (peptides_dict, time_point_list) = remove_blanks_from_list(row, time_point_list, peptides_dict)
            elif rowq_counter == 2:
                data_locations = []
                for position,value in enumerate(row):
                    if value.strip() == 'I0' or value.strip() == 'I1' or value.strip() == 'I2' or value.strip() == 'I3' or value.strip() == 'I4'or value.strip() == 'I5':
                        data_locations.append(position)
                    else:
                        pass
            else:
                row[0] = row[0] + str(rowq_counter)
                peptides_list.append(row[0])
                peptides_dict[row[0]] = []
                data_list = []
                counter_random = 1
                for i in data_locations:
                    if test_for_number(row,i)==True:
                        data_list.append(row[i])
                    else:
                        data_list.append(0)
                    if counter_random == 6:
                        counter_random = 0
                        (I0,I1,I2,I3,I4,I5) = time_names_fxn(data_list)
                        peptides_dict = test_for_zeros(I0,I1,I2,I3,I4,I5,peptides_dict,row)
                        data_list = []
                    else:
                        pass
                    counter_random +=1
            rowq_counter += 1
    if len(peptides_list) == 2:
        pass
    else:
        write_data_main(file_name, peptides_list, time_point_list, peptides_dict, folder_name, time_points, prism_input)

def read_file(folder_name, time_points, prism_input):
    for file_name in os.listdir(folder_name): #folder_name is actually the path to the folder
        file_name_split = file_name.split('.')
        if len(file_name_split) < 2:
            pass
        elif file_name_split[1] == 'Quant':
            file_path = os.path.sep.join([folder_name, file_name])
            with open(file_path) as file:
                rows_counter = 1
                for rowrow in file:
                    
                    if rows_counter == 1:
                        pass
                    elif rows_counter == 2:
                        rowrow = rowrow.split(',')
                        if rowrow[0] == 'Peptide':
                            read_data_A2(file, file_name, file_path, folder_name, time_points, prism_input)
                        else:
                            pass
                    elif rows_counter == 4:
                        rowrow = rowrow.split(',')
                        if rowrow[0] == 'Peptide':
                            read_data_A4(file, file_name, file_path, folder_name, time_points, prism_input)
                        else:
                            pass
                    else:
                        pass
                    rows_counter += 1
        else:
            pass
        

          
def read_file_unique(folder_name, time_points, prism_input):
    for file_name in os.listdir(folder_name): #folder_name is actually the path to the folder
        file_name_split = file_name.split('.')
        if len(file_name_split) < 2:
            pass
        elif file_name_split[1] == 'Quant':
            file_path = os.path.sep.join([folder_name, file_name])
            with open(file_path) as file:
                rows_counter = 1
                for rowrow in file:
                    
                    if rows_counter == 1:
                        pass
                    elif rows_counter == 2:
                        rowrow = rowrow.split(',')
                        if rowrow[0] == 'Peptide':
                            read_data_A2_unique(file, file_name, file_path, folder_name, time_points, prism_input)
                        else:
                            pass
                    elif rows_counter == 4:
                        rowrow = rowrow.split(',')
                        if rowrow[0] == 'Peptide':
                            read_data_A4_unique(file, file_name, file_path, folder_name, time_points, prism_input)
                        else:
                            pass
                    else:
                        pass
                    rows_counter += 1
        else:
            pass
        
def remove_blanks_from_list(row, time_point_list, peptides_dict): #Also adds the time points to the peptides dictionary to make things easier later
    for i in row:
        if 'h' in i or 'H' in i:
            time_point_list.append(i.strip())
        else:
            pass
    peptides_dict['time_points'] = ['']
    for z in time_point_list:
        peptides_dict['time_points'].append(z)
    return (peptides_dict, time_point_list)


def time_names_fxn(data_list):
    I0 = float(data_list[0])
    I1 = float(data_list[1])
    I2 = float(data_list[2])
    I3 = float(data_list[3])
    I4 = float(data_list[4])
    I5 = float(data_list[5])
    return (I0,I1,I2,I3,I4,I5)
                                 
def test_for_number(row,i):#Tests more than just zero
    if row[i].strip() == None:
        return False
    if row[i].strip() == '':
        return False
    else:
        return True

    
def test_for_zeros(I0,I1,I2,I3,I4,I5,peptides_dict,row):
    if I0 + I1 == 0:
        peptides_dict[row[0]].append('')
    elif I1/(I0+I1) == 0 or I1/(I0+I1) > .8 or I1/(I0+I1) < .15: 
        peptides_dict[row[0]].append('')          
    else:
        peptides_dict[row[0]].append(I1/(I0+I1))
        
    if I0 + I1+ I2 == 0:
        peptides_dict[row[0]].append('')
    elif (I1+I2)/(I0+I1+I2) == 0 or (I1+I2)/(I0+I1+I2) == 1 or (I1+I2)/(I0+I1+I2) == 2 or (I1+I2)/(I0+I1+I2) < 0.1: 
        peptides_dict[row[0]].append('') 
    else:
        peptides_dict[row[0]].append((I1+I2)/(I0+I1+I2))
        
    if I0 + I1 + I2 + I3 == 0:
        peptides_dict[row[0]].append('')
    elif (I1+I2+I3)/(I0+I1+I2+I3) == 0 or (I1+I2+I3)/(I0+I1+I2+I3) == 1 or (I1+I2+I3)/(I0+I1+I2+I3) == 2 or (I1+I2+I3)/(I0+I1+I2+I3) < 0.1: 
        peptides_dict[row[0]].append('') 
    else:
        peptides_dict[row[0]].append((I1+I2+I3)/(I0+I1+I2+I3))
        
    if I0 + I1 + I2 + I3 + I4 == 0:
        peptides_dict[row[0]].append('')
    elif (I1+I2+I3+I4)/(I0+I1+I2+I3+I4) == 0 or (I1+I2+I3+I4)/(I0+I1+I2+I3+I4) == 1 or (I1+I2+I3+I4)/(I0+I1+I2+I3+I4) == 2 or (I1+I2+I3+I4)/(I0+I1+I2+I3+I4) < 0.1: 
        peptides_dict[row[0]].append('') 
    else:
        peptides_dict[row[0]].append((I1+I2+I3+I4)/(I0+I1+I2+I3+I4))
        
    if I0 + I1 + I2 + I3 + I4 + I5 == 0:
        peptides_dict[row[0]].append('')
    elif (I1+I2+I3+I4+I5)/(I0+I1+I2+I3+I4+I5) == 0 or (I1+I2+I3+I4+I5)/(I0+I1+I2+I3+I4+I5) == 1 or (I1+I2+I3+I4+I5)/(I0+I1+I2+I3+I4+I5) == 2 or (I1+I2+I3+I4+I5)/(I0+I1+I2+I3+I4+I5) < 0.1: 
        peptides_dict[row[0]].append('') 
    else:
        peptides_dict[row[0]].append((I1+I2+I3+I4+I5)/(I0+I1+I2+I3+I4+I5))

    return peptides_dict

def create_time_points(time_point_list, time_points):
    time_points = time_points.split(',')
    time_points.insert(0,'Times')
    times_list_no_letters = time_points

    return times_list_no_letters
   

def write_data_main(file_name, peptides_list, time_point_list, peptides_dict, folder_name, time_points, prism_input):
    from openpyxl import Workbook
    from openpyxl.compat import range
    times_list_no_letters = create_time_points(time_point_list, time_points)
    
    wb = Workbook()#fix this name
    new_file_name = file_name.split('.')
    new_file_name = new_file_name[0] + '_TL' + '.xlsx'
    new_file_name = os.path.sep.join([folder_name, new_file_name]) 
    
    ws1 = wb.active
    ws1.title = "Total Intensity values"
    
    length_of_peptides_list = len(peptides_list)
    length_of_time_point_list = len(time_point_list)
    counter_for_number_of_tables = 2
    counter_for_rows = 1
    while counter_for_number_of_tables < 7:
        if counter_for_number_of_tables == 2:
            name = 'A' + str(counter_for_rows)#Creates coordinate position of the cell for title of the table
            ws1[name] = 'I1'
            counter_for_value_2 = 0
            counter_for_rows = write_data(peptides_list, length_of_peptides_list, length_of_time_point_list, time_point_list, peptides_dict, counter_for_rows, ws1, counter_for_value_2, times_list_no_letters)
        elif counter_for_number_of_tables == 3:
            name = 'A' + str(counter_for_rows)
            ws1[name] = 'I2'
            counter_for_value_2 = 1
            counter_for_rows = write_data(peptides_list, length_of_peptides_list, length_of_time_point_list, time_point_list, peptides_dict, counter_for_rows, ws1, counter_for_value_2, times_list_no_letters)
        elif counter_for_number_of_tables == 4:
            name = 'A' + str(counter_for_rows)
            ws1[name] = 'I3'
            counter_for_value_2 = 2
            counter_for_rows = write_data(peptides_list, length_of_peptides_list, length_of_time_point_list, time_point_list, peptides_dict, counter_for_rows, ws1, counter_for_value_2, times_list_no_letters)
        elif counter_for_number_of_tables == 5:
            name = 'A' + str(counter_for_rows)
            ws1[name] = 'I4'
            counter_for_value_2 = 3
            counter_for_rows = write_data(peptides_list, length_of_peptides_list, length_of_time_point_list, time_point_list, peptides_dict, counter_for_rows, ws1, counter_for_value_2, times_list_no_letters)
        elif counter_for_number_of_tables == 6:
            name = 'A' + str(counter_for_rows)
            ws1[name] = 'I5'
            counter_for_value_2 = 4
            counter_for_rows = write_data(peptides_list, length_of_peptides_list, length_of_time_point_list, time_point_list, peptides_dict, counter_for_rows, ws1, counter_for_value_2, times_list_no_letters)
        
        counter_for_number_of_tables += 1
    wb.save(filename = new_file_name)
    write_data_prism_main(folder_name, peptides_list, times_list_no_letters, peptides_dict, file_name, prism_input)


    
def write_data(peptides_list, length_of_peptides_list, length_of_time_point_list, time_point_list, peptides_dict, counter_for_rows, ws1, counter_for_value_2, times_list_no_letters):
    from openpyxl.utils import get_column_letter
    counter_for_rows += 1 #counter_for_rows variable tells me which row I am on in the excel file
    counter_for_rows2 = counter_for_rows #Just for formality so I can do the if statement later. Otherwise rowx = counter_for_rows. This just keeps counter_for_rows2 constant until the function ends
    counter_for_time_point_list = 1
    counter_for_peptides_list = 0 #Start at 0 because first row nothing is done and counter adds one too it
    counter_for_times_list_no_letters = 0
    for colx in range(1, length_of_peptides_list+2):
        counter_for_value = 0 + counter_for_value_2
        if colx == 1:
            for rowx in range(counter_for_rows, counter_for_rows + length_of_time_point_list):
                if rowx == counter_for_rows2:
                    cell_name = get_column_letter(colx) + str(rowx)
                    ws1[cell_name] = peptides_list[counter_for_peptides_list]
            
                else:
                    cell_name = get_column_letter(colx) + str(rowx)
                    ws1[cell_name] = time_point_list[counter_for_time_point_list]
                    counter_for_time_point_list += 1
        elif colx == 2:
            counter_for_peptides_list -= 1
            for rowx in range(counter_for_rows, counter_for_rows + length_of_time_point_list):
                    cell_name = get_column_letter(colx) + str(rowx)
                    ws1[cell_name] = times_list_no_letters[counter_for_times_list_no_letters]
                    counter_for_times_list_no_letters += 1
        else:
            for rowx in range(counter_for_rows, counter_for_rows + length_of_time_point_list):
                if rowx == counter_for_rows2:
                    cell_name = get_column_letter(colx) + str(rowx)
                    ws1[cell_name] = peptides_list[counter_for_peptides_list]
                    
                else:
                    cell_name = get_column_letter(colx) + str(rowx)
                    ws1[cell_name] = peptides_dict[peptides_list[counter_for_peptides_list]][counter_for_value]
                    counter_for_value += 5

        counter_for_peptides_list += 1
    counter_for_rows += length_of_time_point_list + 1
    
    #Adding formula for Average and CV
    ws1['A' + str(counter_for_rows)] = '=average(' + 'C' + str(counter_for_rows) + ':' + 'GT' + str(counter_for_rows) + ')' #Adding formula for Average
    ws1['B' + str(counter_for_rows)] = 'K'
    counter_for_rows += 1
    
    ws1['A' + str(counter_for_rows)] = '=stdev(' + 'C' + str(counter_for_rows - 1) + ':' + 'GT' + str(counter_for_rows - 1) + ')' + '/' + 'A' + str(counter_for_rows - 1)
    ws1['B' + str(counter_for_rows)] = 'R^2'
    counter_for_rows +=1

    ws1['B' + str(counter_for_rows)] = 'K'
    counter_for_rows += 1
    return counter_for_rows
    
           
def write_data_prism_main(folder_name, peptides_list, times_list_no_letters, peptides_dict, file_name, prism_input):#Creates csv files for prism
    file_name = file_name.split('.')
    file_name = file_name[0]
    file_name = os.path.sep.join([folder_name, file_name])
    file_name = file_name + '_Prism' + prism_input + '.csv' #Creates file name in correct format
   
    data = create_data_list(peptides_list, times_list_no_letters, peptides_dict, prism_input)
    write_data_prism(data, file_name)
    
    
def create_data_list(peptides_list, times_list_no_letters, peptides_dict, prism_input):
    data = []
    peptides_list.remove('')
    peptides_list.insert(0,'Times')
    data.append(peptides_list)#Added peptides list to data variable

    counter_value = -1 + int(prism_input) #counter to go through time point list to pull out times
    times_list_no_letters.remove('Times')
    counter_times = 0
    for time in times_list_no_letters:
        counter_peptide = 0
        temp_list = []
        temp_list.append(times_list_no_letters[counter_times])
        for peptide in peptides_list:
            if counter_peptide == 0:
                pass
            else:
                temp_list.append(peptides_dict[peptides_list[counter_peptide]][counter_value])
            counter_peptide += 1
        counter_times += 1
        counter_value += 5
        data.append(temp_list)
    return data
def write_data_prism(data, file_name):
    with open(file_name, 'w', newline='') as fp:
        a = csv.writer(fp, delimiter=',')
        a.writerows(data)

def find_prism():
    print('Please wait, looking for Prism Demo.')
    found = 0
    for root,dirs,files in os.walk('C:\\'):
        for file in files:
            if file == 'PrismDemo.exe':
                prism_path = os.path.sep.join([root,file])
                found = 1
                print('Prism Found.')
                break
            else:
                pass
    if found == 0:
        print('Prism 5 Demo was not found in C drive.')
        print('Please place Prism into your C drive.')
        input('The algorithm will now quit. Please press Enter.')
        sys.exit()
    return prism_path

 
def create_prism_script(folder_name, prism_input): #Creates prism script in directory of python script
    cwd = os.getcwd()
    temp = 'SetPath ' + folder_name + '\n'
    path_for_prism_file = 'SetPath ' + cwd + '\n'
    script_info = [path_for_prism_file, 'Open onephase.pzf\n', temp, 'ForEach *Prism' + prism_input +'.csv\n','OpenOutput %F_Results.txt\n','Goto D\n','ClearTable\n','Import\n','Goto R\n','WSheet 5\n','Next\n','Beep']
    with open('prism_script.pzc','w') as file:
        for i in script_info:
            file.write(i)

def run_prism_script(prism_path):
    from subprocess import call
    cwd = os.getcwd()
    script_path = os.path.sep.join([cwd, 'prism_script.pzc'])
    script_path = '@' + script_path
    call([prism_path,script_path])
       
    
def read_file_prism(folder_name, row_counter):
    for file_name3 in os.listdir(folder_name):
        if 'Results.txt' in file_name3:
            file_path = os.path.sep.join([folder_name, file_name3])
            with open(file_path) as file2:
                data_dict = {}
                peptides = []   
                plateau = []
                k_values = []
                k_values_se = []
                half_life = []
                R2 = [] #R^2
                counter_for_line = 1
                hello = 0
                for line in file2:
                    line = line.split('\t')
                    if counter_for_line == 19: #Constraint
                        if line[2] == 'K > 0.0' or line[3] == 'K > 0.0': #R2 is in line 17
                            R2_location = 17
                            break
                        else:
                            pass
                    elif counter_for_line == 20: #Constraint K
                        if line[2] == 'K > 0.0' or line[3] == 'K > 0.0': #R2 is in line 18
                            R2_location = 18
                            break
                        else:
                            pass
                    counter_for_line +=1
                counter_for_line = 1
                file2.seek(0)
                for line in file2:
                    line = line.split('\t')
                    if counter_for_line == 1: #Peptides
                        for i in line:
                            peptides.append(i.strip())
                    elif counter_for_line == 6: #Plateau
                        for i in line:
                            plateau.append(i)                        
                    elif counter_for_line == 7: #K-Values
                        for i in line:
                            k_values.append(i)
                    elif counter_for_line == 9: #Half-time/Half-life
                        for i in line:
                            k_values_se.append(i)
                    elif counter_for_line == 14: #K-values SE
                        for i in line:
                            half_life.append(i)
                    elif counter_for_line == 17 and R2_location == 17:
                        for i in line:
                            R2.append(i)
                    elif counter_for_line == 18 and R2_location == 18:
                        for i in line:
                            R2.append(i)
                    else:
                        pass
                    counter_for_line += 1
        
            for (p,pl,k,kse,h,r) in zip(peptides,plateau,k_values,k_values_se,half_life, R2): #Zip through the lists to create a dictionary of all the data for easy access
                data_dict[p] = [pl,k,kse,h,r] #Dict now created with all the info

            

            create_data_file_individual(data_dict, peptides, file_name3, folder_name) #Create individual data file here
            protein_data, row_counter = prepare_protein_data(file_name3, data_dict, peptides, row_counter)
            protein_data = create_data_file_all(protein_data, folder_name)
        else:
            pass
    
def create_data_file_individual(data_dict, peptides, file_name3, folder_name):
    from openpyxl.utils import get_column_letter
    from openpyxl import Workbook
    new_file_name_3 = file_name3.split('_')
    length_new_file_name_3 = len(new_file_name_3) - 1
    new_file_name_3 = '_'.join(new_file_name_3[0:length_new_file_name_3])
    new_file_name_3 = os.path.sep.join([folder_name, new_file_name_3])
    new_file_name_3 = new_file_name_3 + '_Results.xlsx'
    wby = Workbook()
    wsy = wby.active
    wsy.title = "Data Summary"
    data_names = ['','Plateau','Rate Constant','Half-Life','K-Value SE','R^2']

    for coly in range(1,len(peptides)):
        if coly == 1:
            for rowy in range(1,7):
                cell_namey = get_column_letter(coly) + str(rowy)
                wsy[cell_namey] = data_names[rowy-1]
        else:
            for rowy in range(1,7):
                if rowy == 1:
                    if peptides[coly].find('Data'):
                        cell_namey = get_column_letter(coly) + str(rowy)
                        wsy[cell_namey] = peptides[coly]
                    else:
                        pass                        
                else:
                    if peptides[coly].find('Data'):
                        cell_namey = get_column_letter(coly) + str(rowy)
                        wsy[cell_namey] = data_dict[peptides[coly]][rowy-2]
                    else:
                        pass
                        
                                   
    wby.save(filename = new_file_name_3)


def prepare_protein_data(file_name3, data_dict, peptides, row_counter):
    new_file_name_3 = file_name3.split('_')
    new_file_name_3 = '_'.join(new_file_name_3[0:len(new_file_name_3) - 1])
    protein_data.append([new_file_name_3])#Added protein name
    row_counter +=1

    peptides.insert(2,'CV')
    peptides.insert(2,'Average')
    peptides.insert(0,'')
    peptides = peptides[2:len(peptides)+1]
    temp_peptides = []
    for i in peptides:
        if i.find('Data'): #Function is reverse what you would exepect
            temp_peptides.append(i)
        else:
            pass
    protein_data.append(temp_peptides)#Added peptides
    row_counter +=1
    
    temp_listy = []
    temp_listy.append('K-Value')
    temp_listy.append('=average(' + 'D' + str(row_counter) + ':' + 'GT' + str(row_counter) + ')')
    temp_listy.append('=stdev(' + 'D' + str(row_counter) + ':' + 'GT' + str(row_counter) + ')' + '/' + 'B' + str(row_counter))
    for x in range(3,len(peptides)):
        try:
            if float(data_dict[peptides[x]][4]) < .95 or '~' in data_dict[peptides[x]][1]:
                temp_listy.append('')
            else:
                temp_listy.append(data_dict[peptides[x]][1])
        except ValueError:
            temp_listy.append('')
    protein_data.append(temp_listy)
    row_counter +=1


    temp_listy = []
    temp_listy.append('R^2')
    temp_listy.append('')
    temp_listy.append('')
    for x in range(3,len(peptides)):
        temp_listy.append(data_dict[peptides[x]][4])
    protein_data.append(temp_listy)
    row_counter +=1
    
    
    temp_listy = []
    temp_listy.append('K-Value')
    temp_listy.append('')
    temp_listy.append('')
    for x in range(3,len(peptides)):
        temp_listy.append(data_dict[peptides[x]][1])
    protein_data.append(temp_listy)
    row_counter +=1

    protein_data.append([])
    row_counter +=1
            
    return protein_data, row_counter
    
def create_data_file_all(protein_data, folder_name):
    import csv
    file_path_all = os.path.sep.join([folder_name, 'All_Proteins.csv'])
    with open(file_path_all, 'w', newline = '') as final_file:
        a = csv.writer(final_file, delimiter = ',')
        a.writerows(protein_data)
    return protein_data

def prompt_FSR():
    print('\n')
    print('Please make sure the folder path that contains your data does not contain any spaces. I.e C:\Kasumov lab\Andrew\Data_Folder will not work because there is a space in "Kasumov lab".')
    print('Please enter the path of the folder that contains your files. I.e C:\Kasumov_lab\Andrew\Data_Folder')
    folder_name = input('Folder Path: ').strip()
    folder_name = check_directory(folder_name)

    print('\n')
    print('Please select from the following two choices.')
    print('1. Use M0 and M1 to calculate FSR.')
    print('2. Use M0-M2 to calculate FSR.')
    print('3. Use M0-M3 to calculate FSR.')
    print('4. Use M0-M4 to calculate FSR.')
    print('5. Use M0-M5 to calculate FSR.')
    fsr_selection = input('Enter your choice: ').strip()
    while check_prism_input(fsr_selection)== False:
        print('\n')
        print('This is an invalid input.')
        print('Please try again.')
        print('\n')
        print('Please select from the following two choices.')
        print('1. Use only M0 and M1 to calculate FSR.')
        print('2. Use M0-M2 to calculate FSR.')
        print('3. Use M0-M3 to calculate FSR.')
        print('4. Use M0-M4 to calculate FSR.')
        print('5. Use M0-M5 to calculate FSR.')
        fsr_selection = input('Enter your choice: ').strip()
    
    user_time_points = input('Please enter the timepoints separated by commas without any letters, I.e "0,0,8,8,24,....": ').strip()
    while check_time_points(user_time_points,folder_name) == False:
        print('\n')
        print('The time points you have entered do not match the time points listed in the folder.')
        print('Please try again.')
        user_time_points = input('Please enter the timepoints separated by commas without any letters, I.e "0,0,8,8,24,....": ').strip()

    print('Please enter the body water enrichment for each timepoint as a percent separated by commas, I.e "0,0,2.3,2.4,3.2,....". Make sure to include 0 for each of your controls.')
    bwe = input('Enter the body water enrichment: ').strip()
    while check_time_points(bwe,folder_name) == False:
        print('\n')
        print('The bwe you have entered do not match the time points listed in the folder.')
        print('Please try again.')
        print('Please enter the body water enrichment for each timepoint as a percent separated by commas, I.e "0,0,2.3,2.4,3.2,....". Make sure to include 0 for each of your controls.')
        bwe = input('Enter the body water enrichment: ').strip()
    
    return folder_name, user_time_points, bwe, fsr_selection

def read_file_FSR(folder_name, user_time_points, bwe, fsr_selection):
    for file_name in os.listdir(folder_name): #folder_name is actually the path to the folder
        file_name_split = file_name.split('.')
        if len(file_name_split) < 2:
            pass
        elif file_name_split[1] == 'Quant':
            file_path = os.path.sep.join([folder_name, file_name])
            with open(file_path) as file:
                rows_counter = 1
                for rowrow in file:
                    
                    if rows_counter == 1:
                        pass
                    elif rows_counter == 2:
                        rowrow = rowrow.split(',')
                        if rowrow[0] == 'Peptide':
                            read_data_A2_FSR(file, file_name, file_path, folder_name, user_time_points, bwe, fsr_selection)
                        else:
                            pass
                    elif rows_counter == 4:
                        rowrow = rowrow.split(',')
                        if rowrow[0] == 'Peptide':
                            read_data_A4_FSR(file, file_name, file_path, folder_name, user_time_points, bwe, fsr_selection)
                        else:
                            pass
                    else:
                        pass
                    rows_counter += 1
        else:
            pass

def read_data_A4_FSR(file, file_name, file_path, folder_name, user_time_points, bwe, fsr_selection):
    time_point_list = []
    peptides_list = []
    peptides_dict = {} #Need to add peptides to this
    data_dict = {}
    peptide_locations = []
    rowq_counter = 1 #Counter for which row I am on
    with open(file_path) as file:
        for row in file:
            row = row.split(',')
            if rowq_counter == 1 or rowq_counter == 2:
                pass
            elif rowq_counter == 3:#Making list of Timepoints
                for i in row:
                    if len(i.strip())>2:
                        time_point_list.append(i.strip())
                    else:
                        pass
            elif rowq_counter ==4:
                data_locations = []
                for position,value in enumerate(row):
                    if value.strip() == 'I0' or value.strip() == 'I1' or value.strip() == 'I2' or value.strip() == 'I3' or value.strip() == 'I4' or value.strip() == 'I5' :
                        data_locations.append(position)
                    if value.strip() == 'Charge' or value.strip() == 'SeqMass':
                        peptide_locations.append(position)
                    else:
                        pass
            else: #Extracting data for all of the peptides
                row[0] = row[0] + str(rowq_counter)
                peptides_list.append(row[0].strip())
                data_dict[row[0]] = []
                for i in data_locations:
                    data_dict[row[0]].append(row[i])
                peptides_dict[row[0]] = []
                for i in peptide_locations:
                    peptides_dict[row[0]].append(row[i])
            rowq_counter += 1

        if fsr_selection == '1':
            write_file_FSR_1(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe)
        elif fsr_selection == '2':
            write_file_FSR_2(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe)
        elif fsr_selection == '3':
            write_file_FSR_3(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe)
        elif fsr_selection == '4':
            write_file_FSR_4(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe)
        elif fsr_selection == '5':
            write_file_FSR_5(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe)

def read_data_A2_FSR(file, file_name, file_path, folder_name, user_time_points, bwe, fsr_selection):
    time_point_list = []
    peptides_list = []
    peptides_dict = {} #Need to add peptides to this
    data_dict = {}
    peptide_locations = []
    rowq_counter = 1 #Counter for which row I am on
    with open(file_path) as file:
        for row in file:
            row = row.split(',')
            if rowq_counter == 1:#Making list of Timepoints
                for i in row:
                    if len(i.strip())>2:
                        time_point_list.append(i.strip())
                    else:
                        pass
            elif rowq_counter ==2:
                data_locations = []
                for position,value in enumerate(row):
                    if value.strip() == 'I0' or value.strip() == 'I1' or value.strip() == 'I2' or value.strip() == 'I3' or value.strip() == 'I4' or value.strip() == 'I5':
                        data_locations.append(position)
                    if value.strip() == 'Charge' or value.strip() == 'SeqMass':
                        peptide_locations.append(position)
                    else:
                        pass
            else:
                row[0] = row[0] + str(rowq_counter)
                peptides_list.append(row[0].strip())
                data_dict[row[0]] = []
                for i in data_locations:
                    data_dict[row[0]].append(row[i])
                peptides_dict[row[0]] = []
                for i in peptide_locations:
                    peptides_dict[row[0]].append(row[i])
            rowq_counter += 1
        if fsr_selection == '1':
            write_file_FSR_1(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe)
        elif fsr_selection == '2':
            write_file_FSR_2(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe)
        elif fsr_selection == '3':
            write_file_FSR_3(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe)
        elif fsr_selection == '4':
            write_file_FSR_4(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe)
        elif fsr_selection == '5':
            write_file_FSR_5(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe)


def write_file_FSR_1(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe):
    from openpyxl import Workbook
    from openpyxl.compat import range
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    new_file_name = file_name.split('.')
    new_file_name = new_file_name[0] + '_FSR1' + '.xlsx'
    print('Writing file ', new_file_name)
    new_file_name = os.path.sep.join([folder_name, new_file_name]) 
    ws1 = wb.active
    ws1.title = "FSR"

    heading1 = ['Peptide','Charge','m/z','N']
    heading2 = ['','M0','M1','MPE M1','Net Lab','BWE','k /hr']
    bwe = bwe.split(',')
    user_time_points = user_time_points.split(',')
    
    counter_for_rows = 1
    peptides_counter = 0
    zeros = 0
    for i in user_time_points:
        if i ==0 or i == '0':
            zeros+=1
        else:
            pass
    for chart in range(len(peptides_list)):
        time_point_list_counter = 0
        data_counter = 0
        counter2 = 1
        for row in range(counter_for_rows,counter_for_rows + len(time_point_list)+3):
            if counter2 == 1:
                for column in range(1,len(heading1)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = heading1[column-1]
            elif counter2 == 2:
                temp_list = []
                temp_list.append(peptides_list[peptides_counter])
                N = calculate_N(peptides_list[peptides_counter])
                for i in peptides_dict[peptides_list[peptides_counter]]:
                    temp_list.append(i.strip())
                temp_list.append(N)            
                for column in range(1,len(temp_list)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = temp_list[column-1]
            elif counter2 == 3:
                for column in range(1,len(heading2)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = heading2[column-1]
            elif counter2 > 3 and counter2 < (zeros+4):
                if counter2 == 4:
                    average_location = counter_for_rows
                    temp_list = []
                    temp_list.append(time_point_list[time_point_list_counter])
                    for i in range(data_counter,data_counter +2):
                        try:
                            temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                        except ValueError:
                            temp_list.append('')
                        data_counter+=1
                    data_counter +=4
                    temp_list.append('=if(and(B' + str(counter_for_rows) + '>0,C' + str(counter_for_rows)+ '>0),C' + str(counter_for_rows)+'/(B'+ str(counter_for_rows)+'+C'+str(counter_for_rows)+'),"")')
                    temp_list.append('=average(D'+str(counter_for_rows)+':D'+str(counter_for_rows+ zeros-1)+')')
                    for column in range(1,len(heading2)-1):
                        cell_name = get_column_letter(column) + str(row)
                        ws1[cell_name] = temp_list[column-1]                 
                else:
                    temp_list = []
                    temp_list.append(time_point_list[time_point_list_counter])
                    for i in range(data_counter,data_counter + 2):
                        try:
                            temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                        except ValueError:
                            temp_list.append('')
                        data_counter+=1
                    data_counter += 4
                    temp_list.append('=if(and(B' + str(counter_for_rows) + '>0,C' + str(counter_for_rows)+ '>0),C' + str(counter_for_rows)+'/(B'+ str(counter_for_rows)+'+C'+str(counter_for_rows)+'),"")')
                    for column in range(1,len(heading2)-2):
                        cell_name = get_column_letter(column) + str(row)
                        ws1[cell_name] = temp_list[column-1]                    
                time_point_list_counter += 1
            else:
                temp_list = []
                temp_list.append(time_point_list[time_point_list_counter])
                for i in range(data_counter,data_counter + 2):
                    try:
                        temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                    except ValueError:
                        temp_list.append('')
                    data_counter+=1
                data_counter +=4
                temp_list.append('=if(and(B' + str(counter_for_rows) + '>0,C' + str(counter_for_rows)+ '>0),C' + str(counter_for_rows)+'/(B'+ str(counter_for_rows)+'+C'+str(counter_for_rows)+'),"")')
                temp_list.append('=D'+str(counter_for_rows)+'-E'+str(average_location))
                temp_list.append(float(bwe[time_point_list_counter]))
                temp_list.append('=E'+str(counter_for_rows)+'/F'+str(counter_for_rows)+'*100/'+str(N)+'/'+str(user_time_points[time_point_list_counter]))
                for column in range(1,len(heading2)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = temp_list[column-1]
                time_point_list_counter +=1
                    
            counter2 += 1
            counter_for_rows +=1
        cell_name = get_column_letter(1)+str(counter_for_rows)
        ws1[cell_name] = ''
        peptides_counter +=1
        counter_for_rows +=1
    wb.save(filename = new_file_name)
    
def write_file_FSR_2(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe):
    from openpyxl import Workbook
    from openpyxl.compat import range
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    new_file_name = file_name.split('.')
    new_file_name = new_file_name[0] + '_FSR2' + '.xlsx'
    print('Writing file ', new_file_name)
    new_file_name = os.path.sep.join([folder_name, new_file_name]) 
    ws1 = wb.active
    ws1.title = "FSR"

    heading1 = ['Peptide','Charge','m/z','N']
    heading2 = ['','M0','M1','M2','Sum','M0','M1','M2','Total','Net Lab','BWE','k /hr']
    bwe = bwe.split(',')
    user_time_points = user_time_points.split(',')
    
    counter_for_rows = 1
    peptides_counter = 0
    zeros = 0
    for i in user_time_points:
        if i ==0 or i == '0':
            zeros+=1
        else:
            pass
    for chart in range(len(peptides_list)):
        time_point_list_counter = 0
        data_counter = 0
        counter2 = 1
        for row in range(counter_for_rows,counter_for_rows + len(time_point_list)+3):
            if counter2 == 1:
                for column in range(1,len(heading1)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = heading1[column-1]
            elif counter2 == 2:
                temp_list = []
                temp_list.append(peptides_list[peptides_counter])
                N = calculate_N(peptides_list[peptides_counter])
                for i in peptides_dict[peptides_list[peptides_counter]]:
                    temp_list.append(i.strip())
                temp_list.append(N)            
                for column in range(1,len(temp_list)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = temp_list[column-1]
            elif counter2 == 3:
                for column in range(1,len(heading2)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = heading2[column-1]
            elif counter2 > 3 and counter2 < (zeros+4):
                if counter2 == 4:
                    average_location = counter_for_rows
                    temp_list = []
                    temp_list.append(time_point_list[time_point_list_counter])
                    for i in range(data_counter,data_counter +3):
                        try:
                            temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                        except ValueError:
                            temp_list.append('')
                        data_counter+=1
                    data_counter +=3
                    temp_list.append('=sum(B'+str(counter_for_rows)+':D'+str(counter_for_rows)+ ')')
                    temp_list.append('=B'+str(counter_for_rows)+'/E'+str(counter_for_rows))
                    temp_list.append('=C'+str(counter_for_rows)+'/E'+str(counter_for_rows))
                    temp_list.append('=D'+str(counter_for_rows)+'/E'+str(counter_for_rows))
                    temp_list.append('=G'+str(counter_for_rows)+'+H'+str(counter_for_rows)+'*2')
                    temp_list.append('=average(I'+str(counter_for_rows)+':I'+str(counter_for_rows+ zeros-1)+')')
                    for column in range(1,len(heading2)-1):
                        cell_name = get_column_letter(column) + str(row)
                        ws1[cell_name] = temp_list[column-1]                 
                else:
                    temp_list = []
                    temp_list.append(time_point_list[time_point_list_counter])
                    for i in range(data_counter,data_counter + 3):
                        try:
                            temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                        except ValueError:
                            temp_list.append('')
                        data_counter+=1
                    data_counter +=3
                    temp_list.append('=sum(B'+str(counter_for_rows)+':D'+str(counter_for_rows)+ ')')
                    temp_list.append('=B'+str(counter_for_rows)+'/E'+str(counter_for_rows))
                    temp_list.append('=C'+str(counter_for_rows)+'/E'+str(counter_for_rows))
                    temp_list.append('=D'+str(counter_for_rows)+'/E'+str(counter_for_rows))
                    temp_list.append('=G'+str(counter_for_rows)+'+H'+str(counter_for_rows)+'*2')
                    for column in range(1,len(heading2)-2):
                        cell_name = get_column_letter(column) + str(row)
                        ws1[cell_name] = temp_list[column-1]                    
                time_point_list_counter += 1
            else:
                temp_list = []
                temp_list.append(time_point_list[time_point_list_counter])
                for i in range(data_counter,data_counter + 3):
                    try:
                        temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                    except ValueError:
                        temp_list.append('')
                    data_counter+=1
                data_counter +=3
                temp_list.append('=sum(B'+str(counter_for_rows)+':D'+str(counter_for_rows)+ ')')
                temp_list.append('=B'+str(counter_for_rows)+'/E'+str(counter_for_rows))
                temp_list.append('=C'+str(counter_for_rows)+'/E'+str(counter_for_rows))
                temp_list.append('=D'+str(counter_for_rows)+'/E'+str(counter_for_rows))
                temp_list.append('=G'+str(counter_for_rows)+'+H'+str(counter_for_rows)+'*2')
                temp_list.append('=I'+str(counter_for_rows)+'-J'+str(average_location))
                temp_list.append(float(bwe[time_point_list_counter]))
                temp_list.append('=J'+str(counter_for_rows)+'/K'+str(counter_for_rows)+'*100/'+str(N)+'/'+str(user_time_points[time_point_list_counter]))
                for column in range(1,len(heading2)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = temp_list[column-1]
                time_point_list_counter +=1
                    
            counter2 += 1
            counter_for_rows +=1
        cell_name = get_column_letter(1)+str(counter_for_rows)
        ws1[cell_name] = ''
        peptides_counter +=1
        counter_for_rows +=1
    wb.save(filename = new_file_name)
def write_file_FSR_3(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe):
    from openpyxl import Workbook
    from openpyxl.compat import range
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    new_file_name = file_name.split('.')
    new_file_name = new_file_name[0] + '_FSR3' + '.xlsx'
    print('Writing file ', new_file_name)
    new_file_name = os.path.sep.join([folder_name, new_file_name]) 
    ws1 = wb.active
    ws1.title = "FSR"

    heading1 = ['Peptide','Charge','m/z','N']
    heading2 = ['','M0','M1','M2','M3','Sum','M0','M1','M2','M3','Total','Net Lab','BWE','k /hr']
    bwe = bwe.split(',')
    user_time_points = user_time_points.split(',')
    
    counter_for_rows = 1
    peptides_counter = 0
    zeros = 0
    for i in user_time_points:
        if i ==0 or i == '0':
            zeros+=1
        else:
            pass
    for chart in range(len(peptides_list)):
        time_point_list_counter = 0
        data_counter = 0
        counter2 = 1
        for row in range(counter_for_rows,counter_for_rows + len(time_point_list)+3):
            if counter2 == 1:
                for column in range(1,len(heading1)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = heading1[column-1]
            elif counter2 == 2:
                temp_list = []
                temp_list.append(peptides_list[peptides_counter])
                N = calculate_N(peptides_list[peptides_counter])
                for i in peptides_dict[peptides_list[peptides_counter]]:
                    temp_list.append(i.strip())
                temp_list.append(N)            
                for column in range(1,len(temp_list)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = temp_list[column-1]
            elif counter2 == 3:
                for column in range(1,len(heading2)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = heading2[column-1]
            elif counter2 > 3 and counter2 < (zeros+4):
                if counter2 == 4:
                    average_location = counter_for_rows
                    temp_list = []
                    temp_list.append(time_point_list[time_point_list_counter])
                    for i in range(data_counter,data_counter +4):
                        try:
                            temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                        except ValueError:
                            temp_list.append('')
                        data_counter+=1
                    data_counter +=2
                    temp_list.append('=sum(B'+str(counter_for_rows)+':E'+str(counter_for_rows)+ ')')
                    temp_list.append('=B'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                    temp_list.append('=C'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                    temp_list.append('=D'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                    temp_list.append('=E'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                    temp_list.append('=H'+str(counter_for_rows)+'+I'+str(counter_for_rows)+'*2+J'+str(counter_for_rows)+'*3')
                    temp_list.append('=average(K'+str(counter_for_rows)+':K'+str(counter_for_rows+ zeros-1)+')')
                    for column in range(1,len(heading2)-1):
                        cell_name = get_column_letter(column) + str(row)
                        ws1[cell_name] = temp_list[column-1]                 
                else:
                    temp_list = []
                    temp_list.append(time_point_list[time_point_list_counter])
                    for i in range(data_counter,data_counter + 4):
                        try:
                            temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                        except ValueError:
                            temp_list.append('')
                        data_counter+=1
                    data_counter +=2
                    temp_list.append('=sum(B'+str(counter_for_rows)+':E'+str(counter_for_rows)+ ')')
                    temp_list.append('=B'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                    temp_list.append('=C'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                    temp_list.append('=D'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                    temp_list.append('=E'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                    temp_list.append('=H'+str(counter_for_rows)+'+I'+str(counter_for_rows)+'*2+J'+str(counter_for_rows)+'*3')
                    for column in range(1,len(heading2)-2):
                        cell_name = get_column_letter(column) + str(row)
                        ws1[cell_name] = temp_list[column-1]                    
                time_point_list_counter += 1
            else:
                temp_list = []
                temp_list.append(time_point_list[time_point_list_counter])
                for i in range(data_counter,data_counter + 4):
                    try:
                        temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                    except ValueError:
                        temp_list.append('')
                    data_counter+=1
                data_counter +=2
                temp_list.append('=sum(B'+str(counter_for_rows)+':E'+str(counter_for_rows)+ ')')
                temp_list.append('=B'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                temp_list.append('=C'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                temp_list.append('=D'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                temp_list.append('=E'+str(counter_for_rows)+'/F'+str(counter_for_rows))
                temp_list.append('=H'+str(counter_for_rows)+'+I'+str(counter_for_rows)+'*2+J'+str(counter_for_rows)+'*3')
                temp_list.append('=K'+str(counter_for_rows)+'-L'+str(average_location))
                temp_list.append(float(bwe[time_point_list_counter]))
                temp_list.append('=L'+str(counter_for_rows)+'/M'+str(counter_for_rows)+'*100/'+str(N)+'/'+str(user_time_points[time_point_list_counter]))
                for column in range(1,len(heading2)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = temp_list[column-1]
                time_point_list_counter +=1
                    
            counter2 += 1
            counter_for_rows +=1
        cell_name = get_column_letter(1)+str(counter_for_rows)
        ws1[cell_name] = ''
        peptides_counter +=1
        counter_for_rows +=1
    wb.save(filename = new_file_name)
def write_file_FSR_4(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe):
    from openpyxl import Workbook
    from openpyxl.compat import range
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    new_file_name = file_name.split('.')
    new_file_name = new_file_name[0] + '_FSR4' + '.xlsx'
    print('Writing file ', new_file_name)
    new_file_name = os.path.sep.join([folder_name, new_file_name]) 
    ws1 = wb.active
    ws1.title = "FSR"

    heading1 = ['Peptide','Charge','m/z','N']
    heading2 = ['','M0','M1','M2','M3','M4','Sum','M0','M1','M2','M3','M4','Total','Net Lab','BWE','k /hr']
    bwe = bwe.split(',')
    user_time_points = user_time_points.split(',')
    
    counter_for_rows = 1
    peptides_counter = 0
    zeros = 0
    for i in user_time_points:
        if i ==0 or i == '0':
            zeros+=1
        else:
            pass
    for chart in range(len(peptides_list)):
        time_point_list_counter = 0
        data_counter = 0
        counter2 = 1
        for row in range(counter_for_rows,counter_for_rows + len(time_point_list)+3):
            if counter2 == 1:
                for column in range(1,len(heading1)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = heading1[column-1]
            elif counter2 == 2:
                temp_list = []
                temp_list.append(peptides_list[peptides_counter])
                N = calculate_N(peptides_list[peptides_counter])
                for i in peptides_dict[peptides_list[peptides_counter]]:
                    temp_list.append(i.strip())
                temp_list.append(N)            
                for column in range(1,len(temp_list)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = temp_list[column-1]
            elif counter2 == 3:
                for column in range(1,len(heading2)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = heading2[column-1]
            elif counter2 > 3 and counter2 < (zeros+4):
                if counter2 == 4:
                    average_location = counter_for_rows
                    temp_list = []
                    temp_list.append(time_point_list[time_point_list_counter])
                    for i in range(data_counter,data_counter +5):
                        try:
                            temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                        except ValueError:
                            temp_list.append('')
                        data_counter+=1
                    data_counter +=1
                    temp_list.append('=sum(B'+str(counter_for_rows)+':F'+str(counter_for_rows)+ ')')
                    temp_list.append('=B'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                    temp_list.append('=C'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                    temp_list.append('=D'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                    temp_list.append('=E'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                    temp_list.append('=F'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                    temp_list.append('=I'+str(counter_for_rows)+'+J'+str(counter_for_rows)+'*2+K'+str(counter_for_rows)+'*3+L'+str(counter_for_rows)+'*4')
                    temp_list.append('=average(M'+str(counter_for_rows)+':M'+str(counter_for_rows+ zeros-1)+')')
                    for column in range(1,len(heading2)-1):
                        cell_name = get_column_letter(column) + str(row)
                        ws1[cell_name] = temp_list[column-1]                 
                else:
                    temp_list = []
                    temp_list.append(time_point_list[time_point_list_counter])
                    for i in range(data_counter,data_counter + 5):
                        try:
                            temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                        except ValueError:
                            temp_list.append('')
                        data_counter+=1
                    data_counter +=1
                    temp_list.append('=sum(B'+str(counter_for_rows)+':F'+str(counter_for_rows)+ ')')
                    temp_list.append('=B'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                    temp_list.append('=C'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                    temp_list.append('=D'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                    temp_list.append('=E'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                    temp_list.append('=F'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                    temp_list.append('=I'+str(counter_for_rows)+'+J'+str(counter_for_rows)+'*2+K'+str(counter_for_rows)+'*3+L'+str(counter_for_rows)+'*4')
                    for column in range(1,len(heading2)-2):
                        cell_name = get_column_letter(column) + str(row)
                        ws1[cell_name] = temp_list[column-1]                    
                time_point_list_counter += 1
            else:
                temp_list = []
                temp_list.append(time_point_list[time_point_list_counter])
                for i in range(data_counter,data_counter + 5):
                    try:
                        temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                    except ValueError:
                        temp_list.append('')
                    data_counter+=1
                data_counter +=1
                temp_list.append('=sum(B'+str(counter_for_rows)+':F'+str(counter_for_rows)+ ')')
                temp_list.append('=B'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                temp_list.append('=C'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                temp_list.append('=D'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                temp_list.append('=E'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                temp_list.append('=F'+str(counter_for_rows)+'/G'+str(counter_for_rows))
                temp_list.append('=I'+str(counter_for_rows)+'+J'+str(counter_for_rows)+'*2+K'+str(counter_for_rows)+'*3+L'+str(counter_for_rows)+'*4')
                temp_list.append('=M'+str(counter_for_rows)+'-N'+str(average_location))
                temp_list.append(float(bwe[time_point_list_counter]))
                temp_list.append('=N'+str(counter_for_rows)+'/O'+str(counter_for_rows)+'*100/'+str(N)+'/'+str(user_time_points[time_point_list_counter]))
                for column in range(1,len(heading2)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = temp_list[column-1]
                time_point_list_counter +=1
                    
            counter2 += 1
            counter_for_rows +=1
        cell_name = get_column_letter(1)+str(counter_for_rows)
        ws1[cell_name] = ''
        peptides_counter +=1
        counter_for_rows +=1
    wb.save(filename = new_file_name)
    


def write_file_FSR_5(file_name, folder_name,time_point_list,peptides_list,peptides_dict,data_dict,user_time_points,bwe):
    from openpyxl import Workbook
    from openpyxl.compat import range
    from openpyxl.utils import get_column_letter
    
    wb = Workbook()
    new_file_name = file_name.split('.')
    new_file_name = new_file_name[0] + '_FSR5' + '.xlsx'
    print('Writing file ', new_file_name)
    new_file_name = os.path.sep.join([folder_name, new_file_name]) 
    ws1 = wb.active
    ws1.title = "FSR"

    heading1 = ['Peptide','Charge','m/z','N']
    heading2 = ['','M0','M1','M2','M3','M4','M5','Sum','M0','M1','M2','M3','M4','M5','Total','Net Lab','BWE','k /hr']
    bwe = bwe.split(',')
    user_time_points = user_time_points.split(',')
    
    counter_for_rows = 1
    peptides_counter = 0
    zeros = 0
    for i in user_time_points:
        if i ==0 or i == '0':
            zeros+=1
        else:
            pass
    for chart in range(len(peptides_list)):
        time_point_list_counter = 0
        data_counter = 0
        counter2 = 1
        for row in range(counter_for_rows,counter_for_rows + len(time_point_list)+3):
            if counter2 == 1:
                for column in range(1,len(heading1)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = heading1[column-1]
            elif counter2 == 2:
                temp_list = []
                temp_list.append(peptides_list[peptides_counter])
                N = calculate_N(peptides_list[peptides_counter])
                for i in peptides_dict[peptides_list[peptides_counter]]:
                    temp_list.append(i.strip())
                temp_list.append(N)            
                for column in range(1,len(temp_list)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = temp_list[column-1]
            elif counter2 == 3:
                for column in range(1,len(heading2)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = heading2[column-1]
            elif counter2 > 3 and counter2 < (zeros+4):
                if counter2 == 4:
                    average_location = counter_for_rows
                    temp_list = []
                    temp_list.append(time_point_list[time_point_list_counter])
                    for i in range(data_counter,data_counter +6):
                        try:
                            temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                        except ValueError:
                            temp_list.append('')
                        data_counter+=1
                    temp_list.append('=sum(B'+str(counter_for_rows)+':G'+str(counter_for_rows)+ ')')
                    temp_list.append('=B'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=C'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=D'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=E'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=F'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=G'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=J'+str(counter_for_rows)+'+K'+str(counter_for_rows)+'*2+L'+str(counter_for_rows)+'*3+M'+str(counter_for_rows)+'*4+N'+str(counter_for_rows)+'*5')
                    temp_list.append('=average(O'+str(counter_for_rows)+':O'+str(counter_for_rows+ zeros-1)+')')
                    for column in range(1,len(heading2)-1):
                        cell_name = get_column_letter(column) + str(row)
                        ws1[cell_name] = temp_list[column-1]                 
                else:
                    temp_list = []
                    temp_list.append(time_point_list[time_point_list_counter])
                    for i in range(data_counter,data_counter + 6):
                        try:
                            temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                        except ValueError:
                            temp_list.append('')
                        data_counter+=1
                    temp_list.append('=sum(B'+str(counter_for_rows)+':G'+str(counter_for_rows)+ ')')
                    temp_list.append('=B'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=C'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=D'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=E'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=F'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=G'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                    temp_list.append('=J'+str(counter_for_rows)+'+K'+str(counter_for_rows)+'*2+L'+str(counter_for_rows)+'*3+M'+str(counter_for_rows)+'*4+N'+str(counter_for_rows)+'*5')
                    for column in range(1,len(heading2)-2):
                        cell_name = get_column_letter(column) + str(row)
                        ws1[cell_name] = temp_list[column-1]                    
                time_point_list_counter += 1
            else:
                temp_list = []
                temp_list.append(time_point_list[time_point_list_counter])
                for i in range(data_counter,data_counter + 6):
                    try:
                        temp_list.append(float(data_dict[peptides_list[peptides_counter]][i]))
                    except ValueError:
                        temp_list.append('')
                    data_counter+=1
                temp_list.append('=sum(B'+str(counter_for_rows)+':G'+str(counter_for_rows)+ ')')
                temp_list.append('=B'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                temp_list.append('=C'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                temp_list.append('=D'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                temp_list.append('=E'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                temp_list.append('=F'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                temp_list.append('=G'+str(counter_for_rows)+'/H'+str(counter_for_rows))
                temp_list.append('=J'+str(counter_for_rows)+'+K'+str(counter_for_rows)+'*2+L'+str(counter_for_rows)+'*3+M'+str(counter_for_rows)+'*4+N'+str(counter_for_rows)+'*5')
                temp_list.append('=O'+str(counter_for_rows)+'-P'+str(average_location))
                temp_list.append(float(bwe[time_point_list_counter]))
                temp_list.append('=P'+str(counter_for_rows)+'/Q'+str(counter_for_rows)+'*100/'+str(N)+'/'+str(user_time_points[time_point_list_counter]))
                for column in range(1,len(heading2)+1):
                    cell_name = get_column_letter(column) + str(row)
                    ws1[cell_name] = temp_list[column-1]
                time_point_list_counter +=1
                    
            counter2 += 1
            counter_for_rows +=1
        cell_name = get_column_letter(1)+str(counter_for_rows)
        ws1[cell_name] = ''
        peptides_counter +=1
        counter_for_rows +=1
    wb.save(filename = new_file_name)

def calculate_N(sequence):
    AA = {'A':4,'R':3.43,'N':1.89,'D':1.89,'C':1.62,'E':3.95,'Q':3.95,'G':2.06,'H':2.88,'I':1,'L':0.6,'K':0.54,'M':1.12,'F':0.32,'P':2.59,'S':2.61,'T':0.2,'Y':0.42,'W':0.08,'V':0.56}
    N = 0
    for i in sequence:
        try:
            N += AA[i.upper()]
        except KeyError:
            pass
    N = "{0:.2f}".format(N)
    return N

def main():
    selection = prompt()
    if selection == '1':
        folder_name = prompt_sadygov()
        data_list = read_file_sadygov(folder_name)
        write_file_sadygov(data_list, folder_name)
        print(len(data_list)-1, 'Proteins were found.')
    elif selection == '2':
        folder_name, time_points, unique, prism_input = prompt_tl()
        if unique == 'no':
            file_name = read_file(folder_name, time_points,prism_input)
        elif unique == 'yes':
            file_name = read_file_unique(folder_name, time_points, prism_input)

        create_prism_script(folder_name, prism_input)
        prism_path = find_prism()
        run_prism_script(prism_path)
        read_file_prism(folder_name, row_counter)
        
    elif selection == '3':
        folder_name, time_points, unique = prompt_tl()
        data_list = read_file_sadygov(folder_name)
        write_file_sadygov(data_list, folder_name)
        print(len(data_list)-1, 'Proteins were found.')

        
        if unique == 'no':
            file_name = read_file(folder_name, time_points, prism_input)
        elif unique == 'yes':
            file_name = read_file_unique(folder_name, time_points, prism_input)

        create_prism_script(folder_name, prism_input)
        prism_path = find_prism()
        run_prism_script(prism_path)
        read_file_prism(folder_name, row_counter)

    elif selection == '4':
        folder_name,user_time_points,bwe, fsr_selection = prompt_FSR()
        read_file_FSR(folder_name, user_time_points, bwe, fsr_selection)

main()
print('Done')
